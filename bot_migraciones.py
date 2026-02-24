import matplotlib.pyplot as plt
import threading
import os
from http.server import HTTPServer, BaseHTTPRequestHandler
from openpyxl import load_workbook
from datetime import datetime, timedelta
from telegram import Update
from telegram.ext import ApplicationBuilder, MessageHandler, filters, ContextTypes

# =========================
# CONFIG
# =========================
TOKEN = os.environ.get("BOT_TOKEN")
RENDER_EXTERNAL_URL = os.environ.get("RENDER_EXTERNAL_URL")
PORT = int(os.environ.get("PORT", 10000))
EXCEL_FILE = "./avance.xlsx"

META_ENLACES = 266
META_BALANCEADORES = 266

# Lista de feriados (YYYY-MM-DD)
FERIADOS = [
    "2026-01-01",
    "2026-02-16",
    "2026-03-16",
    "2026-04-10",
    "2026-05-01",
    "2026-09-16",
    "2026-11-16",
    "2026-12-25"
]
FERIADOS = [datetime.strptime(d, "%Y-%m-%d").date() for d in FERIADOS]

# =========================
# MENU
# =========================
MENU = """
📊 MENU MIGRACIONES

1️⃣ Hoy             4️⃣ Semana pasada
2️⃣ Dashboard 5️⃣ Detalle semana
3️⃣ Semana      6️⃣ Gráfica
"""

# =========================
# CONVERTIR FECHA
# =========================
def convertir_fecha(valor):
    if isinstance(valor, datetime):
        return valor
    valor = str(valor).strip()
    formatos = ["%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d-%b-%Y"]
    for formato in formatos:
        try:
            return datetime.strptime(valor, formato)
        except:
            pass
    raise ValueError(f"Formato fecha no soportado: {valor}")

# =========================
# CARGAR DATOS
# =========================
def cargar_datos():
    datos = []
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            if not row[0]:
                continue
            if isinstance(row[0], str) and "semana" in row[0].lower():
                continue

            fecha = convertir_fecha(row[0])

            enlaces_telmex = int(row[2] or 0)
            enlaces_totalplay = int(row[3] or 0)
            bal_telmex = int(row[5] or 0)
            bal_totalplay = int(row[6] or 0)

            datos.append({
                "fecha": fecha,
                "enlaces": enlaces_telmex + enlaces_totalplay,
                "enlaces_telmex": enlaces_telmex,
                "enlaces_totalplay": enlaces_totalplay,
                "balanceadores": bal_telmex + bal_totalplay,
                "bal_telmex": bal_telmex,
                "bal_totalplay": bal_totalplay
            })
        except Exception as e:
            print("Error fila:", row, e)

    wb.close()
    datos.sort(key=lambda x: x["fecha"])
    return datos

# =========================
# BARRA
# =========================
def barra(p):
    llenos = int(p / 5)
    vacios = 20 - llenos
    return "█"*llenos + "░"*vacios

# =========================
# RITMO PROMEDIO TODOS LOS DÍAS HÁBILES
# =========================
def calcular_ritmo():
    datos = cargar_datos()
    if len(datos) < 2:
        return 0,0

    inc_enlaces = []
    inc_bal = []

    for i in range(1, len(datos)):
        fecha = datos[i]["fecha"].date()
        if fecha.weekday() < 5 and fecha not in FERIADOS:  # solo días hábiles
            inc_enlaces.append(datos[i]["enlaces"] - datos[i-1]["enlaces"])
            inc_bal.append(datos[i]["balanceadores"] - datos[i-1]["balanceadores"])

    if not inc_enlaces or not inc_bal:
        return 0, 0

    prom_enlaces = sum(inc_enlaces) / len(inc_enlaces)
    prom_bal = sum(inc_bal) / len(inc_bal)

    return prom_enlaces, prom_bal

# =========================
# HOY (CON PROYECCIÓN)
# =========================
def comando_hoy():
    datos = cargar_datos()
    if not datos:
        return "Sin datos"

    d = datos[-1]
    porc_enlaces = d["enlaces"]/META_ENLACES*100
    porc_bal = d["balanceadores"]/META_BALANCEADORES*100
    ritmo_enlaces, ritmo_bal = calcular_ritmo()

    # calcular días hábiles restantes para proyección
    fecha_actual = datetime.now()
    faltan_enlaces = META_ENLACES - d["enlaces"]
    faltan_bal = META_BALANCEADORES - d["balanceadores"]

    dias_enlaces = 0
    dias_bal = 0
    temp_fecha = fecha_actual
    while faltan_enlaces > 0:
        temp_fecha += timedelta(days=1)
        if temp_fecha.weekday() < 5 and temp_fecha.date() not in FERIADOS:
            faltan_enlaces -= ritmo_enlaces
            dias_enlaces += 1

    temp_fecha = fecha_actual
    while faltan_bal > 0:
        temp_fecha += timedelta(days=1)
        if temp_fecha.weekday() < 5 and temp_fecha.date() not in FERIADOS:
            faltan_bal -= ritmo_bal
            dias_bal += 1

    fecha_enlaces = fecha_actual + timedelta(days=dias_enlaces)
    fecha_bal = fecha_actual + timedelta(days=dias_bal)

    return f"""
📅 {d["fecha"].strftime("%d-%b-%Y")}
━━━━━━━━━━━━━━━
🔹 ENLACES

Total Migrados: {d["enlaces"]}/{META_ENLACES}
📡 Telmex: {d["enlaces_telmex"]}
🌐 Totalplay: {d["enlaces_totalplay"]}

Avance: {porc_enlaces:.2f}%
Ritmo promedio: {ritmo_enlaces:.2f}/día
Proyección fin: {fecha_enlaces.strftime("%d-%b-%Y")}
━━━━━━━━━━━━━━━
⚖️ BALANCEADORES

Total Instalados: {d["balanceadores"]}/{META_BALANCEADORES}
📡 Telmex: {d["bal_telmex"]}
🌐 Totalplay: {d["bal_totalplay"]}

Avance: {porc_bal:.2f}%
Ritmo promedio: {ritmo_bal:.2f}/día
Proyección fin: {fecha_bal.strftime("%d-%b-%Y")}
"""

# =========================
# DASHBOARD
# =========================
def comando_dashboard():
    datos = cargar_datos()
    d = datos[-1]
    porc_enlaces = d["enlaces"]/META_ENLACES*100
    porc_bal = d["balanceadores"]/META_BALANCEADORES*100
    return f"""
📊 DASHBOARD

Fecha: {d["fecha"].strftime("%d-%b-%Y")}

ENLACES MIGRADOS
{d["enlaces"]}/{META_ENLACES}
{barra(porc_enlaces)}
{porc_enlaces:.2f}%
━━━━━━━━━━━━━━━
BALANCEADORES INSTALADOS
{d["balanceadores"]}/{META_BALANCEADORES}
{barra(porc_bal)}
{porc_bal:.2f}%
"""

# =========================
# SEMANA ACTUAL
# =========================
def comando_semana_actual():
    datos = cargar_datos()
    hoy = datetime.now()
    inicio = hoy - timedelta(days=hoy.weekday())
    texto = "📅 SEMANA ACTUAL\n\n"
    encontrados = False
    for d in datos:
        if d["fecha"] >= inicio:
            encontrados = True
            texto += f'{d["fecha"].strftime("%d-%b")}  E:{d["enlaces"]}  B:{d["balanceadores"]}\n'
    if not encontrados:
        return "Sin datos esta semana"
    return texto

# =========================
# SEMANA PASADA
# =========================
def comando_semana_pasada():
    datos = cargar_datos()
    hoy = datetime.now()
    inicio = hoy - timedelta(days=hoy.weekday()+7)
    fin = inicio + timedelta(days=6)
    texto = "📅 SEMANA PASADA\n\n"
    encontrados = False
    for d in datos:
        if inicio <= d["fecha"] <= fin:
            encontrados = True
            texto += f'{d["fecha"].strftime("%d-%b")}  Enlaces:{d["enlaces"]}  Balanceadores:{d["balanceadores"]}\n'
    if not encontrados:
        return "Sin datos semana pasada"
    return texto

# =========================
# DETALLE SEMANA
# =========================
def comando_detalle_semana():
    datos = cargar_datos()
    hoy = datetime.now()
    inicio = hoy - timedelta(days=hoy.weekday())
    semana = [d for d in datos if d["fecha"] >= inicio]
    if len(semana) < 2:
        return "Sin datos suficientes"
    enlaces = semana[-1]["enlaces"] - semana[0]["enlaces"]
    bal = semana[-1]["balanceadores"] - semana[0]["balanceadores"]
    return f"""
📊 DETALLE SEMANA

Enlaces migrados: {enlaces}
Balanceadores instalados: {bal}
"""

# =========================
# GRAFICA
# =========================
def generar_grafica():
    datos = cargar_datos()
    fechas = [d["fecha"] for d in datos]
    enlaces = [d["enlaces"] for d in datos]
    bal = [d["balanceadores"] for d in datos]
    plt.figure()
    plt.plot(fechas,enlaces)
    plt.plot(fechas,bal)
    plt.grid()
    archivo = "grafica.png"
    plt.savefig(archivo)
    plt.close()
    return archivo

# =========================
# TELEGRAM
# =========================
async def responder(update: Update, context: ContextTypes.DEFAULT_TYPE):
    t = update.message.text.lower()
    if t in ["1","hoy"]:
        await update.message.reply_text(comando_hoy())
    elif t in ["2","dashboard"]:
        await update.message.reply_text(comando_dashboard())
    elif t in ["3","semana actual"]:
        await update.message.reply_text(comando_semana_actual())
    elif t in ["4","semana pasada"]:
        await update.message.reply_text(comando_semana_pasada())
    elif t in ["5","detalle semana"]:
        await update.message.reply_text(comando_detalle_semana())
    elif t in ["6","grafica"]:
        archivo = generar_grafica()
        await update.message.reply_photo(photo=open(archivo,"rb"))
    else:
        await update.message.reply_text("Comando no reconocido")
    await update.message.reply_text(MENU)

# =========================
# HEALTHCHECK
# =========================
class HealthCheck(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"OK")

def run_health():
    server = HTTPServer(("0.0.0.0",10000),HealthCheck)
    server.serve_forever()

# =========================
# RUN WEBHOOK
# =========================


TOKEN = os.environ.get("BOT_TOKEN")
RENDER_EXTERNAL_URL = os.environ.get("RENDER_EXTERNAL_URL")
PORT = int(os.environ.get("PORT", 10000))

app = ApplicationBuilder().token(TOKEN).build()

app.add_handler(MessageHandler(filters.TEXT, responder))

print("BOT MIGRACIONES ACTIVO WEBHOOK")

app.run_webhook(
    listen="0.0.0.0",
    port=PORT,
    webhook_url=f"{RENDER_EXTERNAL_URL}/{TOKEN}"
)
